import sys
import re
import openpyxl
import django.core.exceptions as excepts
import CourseFunApp.models as questionModels

# import CourseFunApp.exam_system as exam_sy

char_num_re = re.compile(r"[A-Z]")

def char_to_num(matchobj): 
    return str(ord(matchobj.group(0))-65)


#--------------------------
def check_FillInBlank(sheet, row, desc, check_string):
    if sheet["I" + str(row)].value:
        return True
    else:
        check_string = check_string + "\t第" + str(row) + "行，填空题：没有答案，请检查题目：" + str(desc) + "\n"
        return False


def parse_FillInBlank(sheet, row, desc, lesson):
    key_num = int(sheet["I" + str(row)].value)

    key_list = []
    for idx in range(10, 10 + key_num):
        key_list.append(str(sheet.cell(column=idx, row=row).value))

    key_string = ','.join(key_list)
    # print(desc + ' - ' + key_string)

    return questionModels.FillInBlankQuestion(  
        description = desc,
        sectionID = lesson,
        flag = 3,
        star = 1,
        key = key_string
    )


#--------------------------
def check_TrueOrFalse(sheet, row, desc, check_string):
    if sheet["I" + str(row)].value:
        return True
    else:
        check_string = check_string + "\t第" + str(row) + "行，填空题：没有答案，请检查题目：" + str(desc) + "\n"
        return False


def parse_TrueOrFalse(sheet, row, desc, lesson):
    key = (sheet["I" + str(row)].value == '正确')    
    # print(desc + ' - ' + str(int(key)))
    return questionModels.TrueOrFalseQuestion(  
        description = desc,
        sectionID = lesson,
        flag = 3,
        star = 1,
        key = int(key)
    )


#--------------------------
def check_Choice(sheet, row, desc, check_string):
    idx = 10
    option = sheet.cell(column=idx, row=row).value
    while option:
        idx += 1
        option = sheet.cell(column=idx, row=row).value

    if idx==10:
        check_string = check_string + "\t第" + str(row) + "行，多选题：没有选项！" + str(desc) + "\n"
        return False

    return True


def parse_Choice(sheet, row, desc, lesson):
    key = sheet['I' + str(row)].value
    key = char_num_re.sub(char_to_num, key)
# >>> ord('b')  # convert char to int
# 98
# >>> chr(100)  # convert int to char
# 'd'

    idx = 10
    option_list = []
    option = str(sheet.cell(column=idx, row=row).value)

    while option != "None":
        option_list.append(option)
        idx += 1
        option = str(sheet.cell(column=idx, row=row).value)      
    
    option = "|-|".join(option_list)

    # print(desc + ' - ' + key + ' - ' + option)
    return questionModels.ChoiceQuestion(  
        description = desc,
        sectionID = lesson,
        flag = 3,
        star = 1,
        options = option,
        key = key
    )    


#--------------------------
def check_MultiChoice(sheet, row, desc, check_string):
    return check_Choice(sheet, row, desc, check_string)


def parse_MultiChoice(sheet, row, desc, lesson):
    char_list = list(sheet['I' + str(row)].value.upper())
    char_list.sort()
    key = ','.join(char_list)
    key = char_num_re.sub(char_to_num, key)

    idx = 10
    option_list = []
    option = str(sheet.cell(column=idx, row=row).value)

    while option != "None":
        option_list.append(option)
        idx += 1
        option = str(sheet.cell(column=idx, row=row).value)
    
    option = "|-|".join(option_list)

    # print(desc + ' - ' + key + ' - ' + option)
    return questionModels.MultiChoiceQuestion(  
        description = desc,
        sectionID = lesson,
        flag = 3,
        star = 1,
        options = option,
        key = key
    )
    

#--------------------------
def check_Pair(sheet, row, desc, check_string):
    idx = 0
    option_l = sheet.cell(column=10+idx<<1, row=row).value
    option_r = sheet.cell(column=11+idx<<1, row=row).value
    
    while option_l:
        idx += 1
        option_l = sheet.cell(column=10+idx<<1, row=row).value
        option_r = sheet.cell(column=11+idx<<1, row=row).value

    if idx==0:
        check_string = check_string + "\t第" + str(row) + "行，配对题：没有选项！" + str(desc) + "\n"
        return False        
    return True


def parse_Pair(sheet, row, desc, lesson):
    idx = 0
    option_l_list = []
    option_r_list = []
    option_l = str(sheet.cell(column=10+idx<<1, row=row).value)
    option_r = str(sheet.cell(column=11+idx<<1, row=row).value)

    while option_l != "None":
        option_l_list.append(option_l)
        option_r_list.append(option_r)
        idx += 1
        option_l = str(sheet.cell(column=10+idx<<1, row=row).value)
        option_r = str(sheet.cell(column=11+idx<<1, row=row).value)
    
    option_l = "|-|".join(option_l_list)
    option_r = "|-|".join(option_r_list)

    return questionModels.PairQuestion(  
        description = desc,
        sectionID = lesson,
        flag = 3,
        star = 1, 
        leftOptions = option_l,
        rightOptions = option_r 
    )


#--------------------------
def check_Sort(sheet, row, desc, check_string):
    idx = 10
    option = sheet.cell(column=idx, row=row).value
    while option:
        idx += 1
        option = sheet.cell(column=idx, row=row).value

    if idx==10:
        check_string = check_string + "\t第" + str(row) + "行，排序题：没有选项！" + str(desc) + "\n"
        return False

    return True


def parse_Sort(sheet, row, desc, lesson):
    idx = 10
    option_list = []
    option = str(sheet.cell(column=idx, row=row).value)

    while option != "None":
        option_list.append(option)
        idx += 1
        option = str(sheet.cell(column=idx, row=row).value)        
    
    option = "|-|".join(option_list)
    # print(desc + ' - ' + option)

    return questionModels.SortQuestion(  
        description = desc,
        sectionID = lesson,
        flag = 3,
        star = 1,
        options = option,
    )


#--------------------------
def check_Subject(sheet, row, desc, check_string):
    return True


def parse_Subject(sheet, row, desc, lesson):
    # print(desc) 
    return questionModels.CaseAnalyseQuestion(  
        description = desc,
        sectionID = lesson,
        flag = 3,
        star = 1,
    )       


#--------------------------
def update_DB_from_excel(excel_url, db_name):
    lesson_list = questionModels.Lesson.objects.using(db_name).all()

    wb = openpyxl.load_workbook(filename=excel_url, read_only=True, data_only=True)
    # sectionID
    # description
    config_sht = wb['config']
    
    type_dict = {}
    for row in range(2, 8):
        qTypeName = str(config_sht.cell(row=row, column=4).value)
        type_dict[qTypeName] = str(config_sht.cell(row=row, column=5).value)

    for row in range(2, 8):
        qTypeName = config_sht.cell(row=row, column=4).value

        # print("----" + str(qTypeName))
        cur_sht = wb[qTypeName]
        if cur_sht: 
            row_idx = 3
            ques_desc = str(cur_sht['G' + str(row_idx)].value)

            while ques_desc != "None":
                try:
                    ques_type = cur_sht['F' + str(row_idx)].value
                    lesson_desc = cur_sht['A' + str(row_idx)].value                        
                    lesson = lesson_list.get(description=lesson_desc)
                    
                    parsefunc = getattr(sys.modules[__name__], "parse_" + ques_type)
                    quest = parsefunc(cur_sht, row_idx, ques_desc, lesson)                    
                    quest.save(using=db_name)
                    # print(ques_type, quest)
                except (AttributeError) as e:
                    print("---- " + e, ques_desc, lesson_desc, row_idx, ques_type)
                except (excepts.ObjectDoesNotExist) as e:
                    print("??? ", str(e), ques_desc, lesson_desc, row_idx, ques_type)

                row_idx += 1
                ques_desc = str(cur_sht['G' + str(row_idx)].value)


def import_lesson_list(excel_url, db_name):
    wb = openpyxl.load_workbook(filename=excel_url, read_only=True, data_only=True)

    questionModels.Lesson.objects.using(db_name).all().delete()
    lesson_list_sht = wb['课程目录']
    if lesson_list_sht:
        row_idx = 2
        lesson_desc = lesson_list_sht['A' + str(row_idx)].value
        while lesson_desc:
            try:
                questionModels.Lesson.objects.using(db_name).create(description=lesson_desc)
            except (Exception) as e:
                print("---- " + str(e))               
            row_idx += 1
            lesson_desc = lesson_list_sht['A' + str(row_idx)].value


def check_question_excel(excel_url, db_name):
    valid = True
    check_string = "文件检查结果：\n"
    try:
        wb = openpyxl.load_workbook(filename=excel_url, read_only=True, data_only=True)
    except (Exception) as e:
        check_string = check_string + "Excel文件版本太旧！\n" + str(e)
        return False, check_string

    check_sheet = wb['课程目录']
    if check_sheet:
        row_idx = 2
        lesson_desc = check_sheet['A' + str(row_idx)].value
        while lesson_desc:
            row_idx += 1
            lesson_desc = check_sheet['A' + str(row_idx)].value
        check_string = check_string + "<课程目录>共" + str(row_idx-2) + "章\n\n"
        valid = True and valid
    else:                     
        check_string = check_string + "注意：<课程目录>页缺失\n\n"
        valid = False

    config_sht = wb['config']
    if config_sht:
        type_dict = {}
        for row in range(2, 8):
            qTypeName = str(config_sht.cell(row=row, column=4).value)
            type_dict[qTypeName] = str(config_sht.cell(row=row, column=5).value)

        lesson_list = questionModels.Lesson.objects.using(db_name).all()
        for row in range(2, 8):
            qTypeName = str(config_sht.cell(row=row, column=4).value)
            check_sheet = wb[qTypeName]
            if check_sheet: 
                check_string = check_string + "题型<" + qTypeName + ">页面检查结果：\n"
                row_idx = 3
                ques_desc = str(check_sheet['G' + str(row_idx)].value)
                while ques_desc != "None":
                    if check_sheet['F' + str(row_idx)].value:
                        ques_type = str(check_sheet['F' + str(row_idx)].value)
                    else:
                        ques_type = type_dict[str(check_sheet['E' + str(row_idx)].value)]
                        
                    lesson_desc = check_sheet['A' + str(row_idx)].value

                    if lesson_list.filter(description=lesson_desc).exists():
                        try:    
                            checkfunc = getattr(sys.modules[__name__], "check_" + ques_type)
                            valid = valid and checkfunc(check_sheet, row_idx, ques_desc, check_string)
                        except (AttributeError) as e:
                            check_string = check_string + "\t第" + str(row_idx) + "行 |题目类型|：<" + str(ques_type) + "> 错误! 错误信息：" + str(e) + "\n"
                            valid = False
                        except (Exception) as e:
                            print(str(e))
                            check_string = check_string + "\t第" + str(row_idx) + "行 <" + str(ques_type) + ">题，题目有错  错误信息：" + str(e) + "\n"
                            valid = False                            
                    else:
                        check_string = check_string + "\t第" + str(row_idx) + "行 第一列，所属单元:" + str(lesson_desc) + " 错误或者不存在!\n"
                        valid = False

                    row_idx += 1
                    ques_desc = str(check_sheet['G' + str(row_idx)].value)
            else:
                check_string = check_string + "题型<" + qTypeName + ">页面缺失\n"
                valid = False
    else:
        check_string = check_string + "注意：配置文件页面缺失，页面名称<config>\n\n"
        valid = False

    return valid, check_string
    

# database_tool.
# update_DB_from_excel('D:/Temp/DB_1.xlsm')